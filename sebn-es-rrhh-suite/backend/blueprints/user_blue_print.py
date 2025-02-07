import re
from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from config.config import token_required
from config.config import Config
import utils.database as data_utils
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

user_blue_print = Blueprint('user_blue_print', __name__)
from blueprints.travel_blue_print import format_travel_details_json

@user_blue_print.route('/get-bellow-employees', methods=["GET"])
@token_required
def get_bellow_employees_responsible(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            employees = database.get_users_below(current_user)
            for employee in employees:
                employee_data = [{'FullName': employee["FullName"], 'Email': employee["Email"], 'TechResponsible': employee["TechResponsible"], 'TravelResponsible': employee["TravelResponsible"]} for employee in employees]
        
            for employee in employee_data: 
                employee_email = employee['Email']
                if(employee['TechResponsible'] == current_user):
                    user_vacations_result = get_holidays(employee_email)
                    employee['Vacations'] = user_vacations_result
                
                if(employee['TravelResponsible'] == current_user):
                    user_travels_raw =  database.select_travels_by_user(employee_email)
                    user_travels_result = [format_travel_details_json(travel) for travel in user_travels_raw]
                    employee['Travels'] = user_travels_result

        return jsonify({'Employees': employee_data}), 200
    except UnboundLocalError:
            return jsonify({'state': 'Current user is not Tech Responsible'}), 400
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    

#TODO: test_get_responsible_has_unsigned
@user_blue_print.route('/get-responsible-has-unsigned', methods=["GET"])
@token_required
def get_responsible_has_unsigned(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            employees_unsigned = database.get_responsible_has_unsigned(current_user)
        return jsonify(employees_unsigned), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500 
    
    
@user_blue_print.route('/get-user-logs', methods=["GET"])
@token_required
def get_logs(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            logs = database.get_logs()
            logs_json = [
                {'id': log[0], 'UserEmail': log[1], 'EntryDate': log[2], 'StartingDay': log[3], 'FinishingDay': log[4], 'Status': log[6], 'Quantity': log[7]} for log in logs]
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    return jsonify(logs_json), 200


#TODO: shouldn't this function be in hr_blue_print????
""" To add inputs to a user in hr/inputs """
@user_blue_print.route('/update-user-single-day', methods=["POST"])
@token_required
def add_input(current_user):
    data = request.get_json()
    email = data.get('email') 
    dayType = data.get('type')
    quantity = data.get('quantity')
    reason = data.get('reason')

    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            user_days = database.get_single_user_days(dayType, email) 

            if user_days[0][0] + quantity < 0:
                return jsonify({'state': 'Cannot subtract days, resulting quantity would be less than zero'}), 400

            database.create_single_user_days(dayType, quantity, email)
            database.create_hr_logs(email, datetime.now(), reason, dayType, quantity)
        return jsonify({'state': 'SUCCESS'}), 200
    except data_utils.pymssql.Error as e:
        return jsonify({'state': str(e)}), 500


@user_blue_print.route('/get-user-holidays', methods=["POST"])
@token_required
def get_user_vacations(current_user):
    data = request.get_json() 
    email = data.get('email')
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            vacations = database.get_user_holidays(email)  
            vacations_json = [
                {'id': vacation[0], 'startingDay': vacation[3].strftime('%Y-%m-%d %H:%M'), 'finishingDay': vacation[4].strftime('%Y-%m-%d %H:%M'), 'partialDay':vacation[5], 'status': vacation[6], 'totalQuantity': vacation[7]} for vacation in vacations]
            return jsonify(vacations_json), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'error'}), 500


@user_blue_print.route('/create-user-holidays-excel', methods=["GET"])
@token_required
def export_vacations_excel(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            holidays = database.get_user_holidays(current_user)
        
        print(holidays)
        holiday_records = [
            {
                'Start': holiday[3].date(),
                'Finish': holiday[4].date(),
                'Status': holiday[6],
                'PartialDay': holiday[5],
                'TotalQuantity': holiday[7],
                'EntryDate': holiday[2]
            }
            for holiday in holidays
        ]

        df = pd.DataFrame(holiday_records)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='User Data', index=False, startrow=3)

            worksheet = writer.sheets['User Data']

            worksheet['A2'] = 'User'
            worksheet['B2'] = current_user

            worksheet['A4'] = 'Start'
            worksheet['B4'] = 'Finish'
            worksheet['C4'] = 'Status'
            worksheet['D4'] = 'PartialDay'
            worksheet['E4'] = 'TotalQuantity'
            worksheet['F4'] = 'EntryDate'

            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)

            for cell in worksheet["4:4"]:
                cell.fill = header_fill
                cell.font = header_font

            worksheet.column_dimensions["A"].width = 12
            worksheet.column_dimensions["B"].width = 12
            worksheet.column_dimensions["C"].width = 12
            worksheet.column_dimensions["D"].width = 12
            worksheet.column_dimensions["E"].width = 15
            worksheet.column_dimensions["F"].width = 15

            tab = Table(displayName="UserHolidays", ref=f"A4:F{len(df) + 4}")

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style

            worksheet.add_table(tab)

        output.seek(0)

        # Preparar la respuesta para enviar el archivo Excel
        today_str = datetime.today().strftime('%Y-%m-%d')
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'{current_user}_data_{today_str}.xlsx'), 200

    except data_utils.pymssql.Error as e:
        return jsonify({'state': 'DB error:' + str(e)}), 500


def get_holidays(email):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            holidays = database.get_user_holidays(email)
            holidays_json = [
                {'id': holiday[0], 'startingDay': holiday[3], 'finishingDay': holiday[4], 'partialDay':holiday[5], 'status': holiday[6], 'totalQuantity': holiday[7]} for holiday in holidays]
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'})
    return holidays_json


@user_blue_print.route('/get-user-city', methods=["GET"])
@token_required
def get_user_city(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            user_info = database.get_user_info(current_user)[0]
        city = user_info[7]
        return jsonify({'City': city})
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'})


@user_blue_print.route('/get-department-partners', methods=["GET"])
@token_required
def get_department_partners(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            users = database.get_department_partners(current_user)
            users_json = [{'FullName': user[0]} for user in users]
            return jsonify(users_json)
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'})
    
    
@user_blue_print.route('/get-user-details', methods=["POST"])
@token_required
def get_user_details(current_user):
    data = request.get_json() 
    email = data.get('email')
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            user_data = database.fetch_all_user_data(email)
            users_json = {'FullName': user_data[0], 'Email': user_data[1], 'Department': user_data[2], 'Position': user_data[3], 'SubDepartment': user_data[4], 'TechResponsible': user_data[5], 'TravelResponsible': user_data[6], 'City': user_data[7], 'CostCenter': user_data[12], 'WorkingHours': user_data[13]} 
            return jsonify(users_json), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500


@user_blue_print.route('/update-user-details', methods=["PUT"])
@token_required
def update_user_details(current_user):
    data = request.get_json() 
    email = data.get('email')
    full_name = data.get('fullName')
    city = data.get('city')
    position = data.get('position')
    department = data.get('department') 
    sub_department = data.get('subDepartment') 
    cost_center = data.get('costCenter') 
    working_hours = data.get('workingHours') 
    tech_responsible = data.get('techResponsible') 
    travel_responsible = data.get('travelResponsible') 
    
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            database.update_user_data(full_name, email, city, position, department, sub_department, cost_center, working_hours, tech_responsible, travel_responsible)           
    except data_utils.pymssql.Error as ex:
        return jsonify({'state': 'DB error: ' + str(ex)}), 500
    return jsonify({'state':'SUCCESS'}), 200
   
   
@user_blue_print.route('/get-user-cost-center', methods=["GET"])
@token_required
def get_user_cost_center(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            cost_center = database.select_user_cost_center(current_user)
            return jsonify(cost_center), 200
    except data_utils.pymssql.Error as ex:
        return jsonify({'state': 'DB error: '+ str(ex)}), 500
    
    
#TODO: test_get_user_working_hours
@user_blue_print.route('/get-user-working-hours', methods=["GET"])
@token_required
def get_user_working_hours(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            cost_center = database.select_user_working_hours(current_user)
            return jsonify(cost_center), 200
    except data_utils.pymssql.Error as ex:
        return jsonify({'state': 'DB error: '+ str(ex)}), 500
      
   
@user_blue_print.route('/delete-user', methods=["PUT"])
@token_required
def delete_user(current_user):
    data = request.get_json() 
    email = data.get('email')
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            database.delete_user(email)
            return jsonify({'state': 'SUCCESS'}), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500   
    
    
""" CAREFUL, THIS ERASES USER FROM DATABASE FOR REAL """
@user_blue_print.route('/delete-user-database', methods=["DELETE"])
@token_required
def delete_user_from_database(current_user):
    data = request.get_json() 
    email = data.get('email')
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            database.delete_user_from_database(email)
            return jsonify({'state': 'SUCCESS'}), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500   
    
    
#TODO: test_create_user_info_report
@user_blue_print.route('/get-user-info-report', methods=['POST'])
@token_required
def create_user_info_report(create_user):
    data = request.get_json()
    user_info = data.get('userData')
    
    details:dict = user_info['userDetails']
    buckets:dict = user_info['userBuckets']
    home_office:dict = user_info['userHomeOffice']
    vacations:dict = user_info['userVacations']
    travels:dict = user_info['userTravels']
    
    try:
        sheet_name = details['FullName']
        wb = Workbook()
        #remove default sheet and create new one
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
        wb.create_sheet(title=sheet_name, index=0)
        ws = wb[sheet_name]
        
        start_row_titles = 1
        start_row_headers = 2
        start_row_data = 3
        
        start_col_details = 1
        start_col_buckets = 12
        start_col_home_office = 15
        start_col_vacations = 18
        start_col_travels = 25
        
        fill_details_header = PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid")
        fill_details_data = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        fill_buckets_header = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        fill_buckets_data = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        fill_home_office_header = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
        fill_home_office_data = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        fill_vacations_header = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        fill_vacations_data = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        fill_travels_header = PatternFill(start_color="BC9BFF", end_color="BC9BFF", fill_type="solid")
        fill_travels_data = PatternFill(start_color="D3BDFF", end_color="D3BDFF", fill_type="solid")
        
        #DETAILS TABLE: (this table is built differently bc data is returned in a diff format :c)
            # details_headers = ["FullName","Email","Position","Department","SubDepartment","CostCenter","WorkingHours","City","TechResponsible","TravelResponsible"]
            # create_excel_table(ws, details, details_headers, "Details", start_col_details, fill_details_header, fill_details_data, start_row_headers, start_row_data)
        row = start_row_headers
        col = start_col_details
        #title
        end_col = start_col_buckets-2
        build_excel_table_title(ws,"User Details",start_col_details,end_col,fill_details_header)
        #table
        for header, value in details.items():
            header_formatted = format_table_header_name(header)
            #autoadjust the col width automatically
            col_width = max(len(str(header_formatted)), len(str(value))) + 2 #2 for extra margin
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = col_width
            
            cur_cell = ws.cell(row=row, column=col, value=header_formatted.title())
            cur_cell.fill = fill_details_header
            cur_cell.font = Font(bold=True)
            cur_cell.alignment = Alignment(horizontal='justify', vertical='center') 
            
            cur_cell = ws.cell(row=row+1, column=col, value=value)
            cur_cell.fill = fill_details_data
            cur_cell.alignment = Alignment(horizontal='justify', vertical='center') 
            col+=1
        
        #BUCKETS TABLE:
        buckets_headers = ["BucketName","BucketValue"]
        create_excel_table(ws, buckets, buckets_headers, "Buckets", start_col_buckets, fill_buckets_header, fill_buckets_data, start_row_headers, start_row_data)
            
        #HOME OFFICE TABLE:
        home_office_headers = ["Date","Type"]
        create_excel_table(ws, home_office, home_office_headers, "Home Office", start_col_home_office, fill_home_office_header, fill_home_office_data, start_row_headers, start_row_data)
            
        #VACATIONS TABLE:
        vacations_headers = ["id","startingDay","finishingDay","partialDay","totalQuantity","status"]
        create_excel_table(ws, vacations, vacations_headers, "Holidays", start_col_vacations, fill_vacations_header, fill_vacations_data, start_row_headers, start_row_data)
        
        #TRAVELS TABLE:
        travel_headers = ["id", "dateSubmitted", "dateFrom", "dateTo", "dayAmount", "reason", "destination", "address", "costCenter", "addedValue", "advancePayment", "advancePaymentAmount", "transport", "extraDocumentation", "responsibleSignature", "managementSignature", "status"]
        create_excel_table(ws, travels, travel_headers, "Travels", start_col_travels, fill_travels_header, fill_travels_data, start_row_headers, start_row_data)

        # Send the file as a response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"UserInfo_{sheet_name}_{datetime.now().date()}.xlsx"
        return send_file(output, download_name=file_name, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'), 200
    except Exception as ex:
        return jsonify({'state': 'Error: ' + str(ex)}), 500
    
def format_table_header_name(header_text):
        return re.sub(r'([A-Z])', r' \1', header_text).title()
    
def build_excel_table_title(ws, table_title, start_col, end_col, fill_color):
    title_cell = ws.cell(row=1, column=start_col, value=table_title)   
    title_cell.alignment = Alignment(horizontal='center', vertical='center') 
    title_cell.font = Font(bold=True,size=14)
    title_cell.fill = fill_color
    title_details_range = f'{get_column_letter(start_col)}1:{get_column_letter(end_col)}1'
    ws.merge_cells(title_details_range)
    
def create_excel_table(ws, array, table_headers_list, table_title, start_col, fill_color_header, fill_color_data, start_row_headers, start_row_data):
    col = start_col
    col_widths = []
    #TODO: how tf to set cell datatype???????
    cell_data_types = {"<class 'int'>":"'n'", "<class 'str'>":"'s'", "<class 'datetime.date'>":"'d'", "<class 'bool'>":"'b'", "<class 'NoneType'>":"'s'"} #string = 's', date='d', bool='b', number='n'
    
    #title:
    end_col = start_col + len(table_headers_list)-1
    build_excel_table_title(ws,table_title,start_col,end_col,fill_color_header)
    
    #headers:
    for i in range(len(table_headers_list)):
        header_value = format_table_header_name(table_headers_list[i])
        header = ws.cell(row=start_row_headers, column=col+i, value=header_value)
        header.fill = fill_color_header
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal='justify', vertical='center') 
        
        #autoadjust col width (we set default col widths)
        column_letter = get_column_letter(col+i)
        ws.column_dimensions[column_letter].width = len(str(header_value)) + 3
        col_widths.insert(i, len(str(header_value)) + 3)
    
    #rows: 
    row = start_row_data
    for row_data in array:
        for i in range(len(table_headers_list)):
            cell_value = row_data[table_headers_list[i]] if row_data[table_headers_list[i]] != None else "---"
            cur_cell = ws.cell(row=row, column=col+i, value=cell_value)
            # cur_cell.data_type = cell_data_types[str(type(cell_value))] #this does nothing
            cur_cell.fill = fill_color_data
            cur_cell.alignment = Alignment(horizontal='justify', vertical='center') 
            
            #autoadjust the col width automatically (we check if new cell is wider than its col width)
            column_letter = get_column_letter(col+i)
            col_width = max(len(str(cell_value))+2, col_widths[i]) #2 for extra margin
            col_widths[i] = col_width
            ws.column_dimensions[column_letter].width = col_width
        row+=1  