from decimal import Decimal
from flask import Blueprint, jsonify, request, send_file
from config.config import Config, token_required
import utils.database as data_utils
import utils.Hiring_RRHH_CUE as hiring
from utils.mailing_system import send_email
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import string
import random
import logging
from io import BytesIO

hr_blue_print = Blueprint('hr_blue_print', __name__)

@hr_blue_print.route('/get-hr-logs', methods=["GET"])
@token_required
def get_hr_logs(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            logs = database.get_hr_logs()
            print(logs)
            logs_json = [{'id': log[0], 'UserEmail': log[1], 'EntryDate': log[2], 'Reason': log[3], 'Bucket': log[4], 'Quantity': log[5]} for log in logs]
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    return jsonify(logs_json), 200


@hr_blue_print.route('/get-departments', methods=["GET"])
@token_required
def get_departments(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            departments = database.get_departments()
            departments_json = [{'name': d[0], 'Email': d[1]} for d in departments]
            return jsonify(departments_json), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'An error occurred while obtaining the departments'}), 500


@hr_blue_print.route('/get-sub-departments', methods=["GET"])
@token_required
def get_sub_departments(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            sub_departments = database.get_sub_departments()
            sub_departments_json = [{'SubDepartment': data[0], 'SubDepartmentEmail': data[1], 'Department': data[2], 'Responsible': data[3]} for data in sub_departments]
            return jsonify(sub_departments_json), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'An error occurred while obtaining the subDepartments'}), 500


@hr_blue_print.route('/get-buckets', methods=["GET"])
@token_required
def get_buckets(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            buckets = database.get_buckets()
            buckets_json = [{'bucket': b[0]} for b in buckets]
    except data_utils.pymssql.Error as e:
        return jsonify({'state': 'DB ERROR: '+e}), 500
    return jsonify(buckets_json), 200


@hr_blue_print.route('/create-user', methods=["POST"])
@token_required
def create_user(current_user):
    data = request.get_json()

    required_fields = ['fullname','email','department','position','techResponsible','travelResponsible','city','gerency','startDate']
    if not all(data.get(field) and data.get(field).strip() for field in required_fields):
        return jsonify({'state': 'All fields are required'}), 400
    
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            users_data = database.get_all_existing_users() #we also check users that arent working anymore but are still registered in the database
            
        users_json:dict = {}
        for user in users_data: users_json.update({f"{user[1]}":f"{user[0]}"}) 
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500

    if data.get('email') in users_json.keys():
        return jsonify({'state':'A user with that email already exists'}), 400

    fullname = data.get('fullname')
    if len(fullname.split()) < 2:
        return jsonify({'state': 'The full name must contain at least a first and a last name'}), 400

    try:
        start_date = datetime.strptime(data.get('startDate'),'%Y-%m-%d').date()

        gerency = data.get('gerency')
        email = data.get('email')

        void_password = string_generator()
        hashed_password, salt = data_utils.encode_password(void_password)

        holidays = user_days_calculation(data['city'], start_date)

        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            database.create_user(fullname, email, data['department'], data['subDepartment'], data['costCenter'], data['workingHours'], data['position'], data['techResponsible'],
                                 data['travelResponsible'], data['city'], hashed_password, salt)
            database.create_user_days(email, holidays)
            database.add_employee_to_manager(email, gerency)
        
        send_email(email, 'Welcome to SEBN ES SUITE!',  f"<p>Your user has successfully been registred in our system.</p>Your email: {email}<br/>Your password: {void_password} <br/><br/>We recommend that you change your password as soon as you log in and, please, <b>do NOT share this information</b> with anyone.</br>Thank you.")
    except Exception as ex:
        return jsonify({'state': 'Error: ' + str(ex)}), 500
    return jsonify({'state': 'SUCCESS'}), 200


def user_days_calculation(city, start_date):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            city_days = database.select_city_days()
        all_city_days = [{city_day[1]: city_day[2]} for city_day in city_days]
        
        #build city vacation days json
        city_vacation_days = {}
        for city_day in all_city_days:
            city_vacation_days.update(city_day)
        
        #calc user vacation days
        city_value = city_vacation_days[city]
        end_of_year = datetime(datetime.today().year, 12, 31)
        start_date_datetime = datetime.combine(start_date, datetime.min.time())
        days_left = (end_of_year - start_date_datetime).days
        return (days_left * city_value) / 365
    
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    except Exception as ex:
        return jsonify({'state': 'Error: ' + str(ex)}), 500

    
#TODO: test_get_city_vacation_days  
@hr_blue_print.route('/get-city-days', methods=["GET"])
@token_required
def get_city_vacation_days(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            vacations = database.select_city_days()
            all_vacations = [{"Year":vacation[0], "City": vacation[1], "Days":vacation[2]} for vacation in vacations]
            
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    return jsonify(all_vacations), 200
    

def string_generator():
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for _ in range(8))


"""NOT USED"""
#overwrites the exiting values
@hr_blue_print.route('/edit-user-days', methods=["PUT"])
@token_required
def edit_user_days(current_user):
    data = request.get_json()
    email = data.get('Email')
    user_days = data.get('UserDays')
    extra_hours = user_days[0].get('ExtraHours')
    holidays = user_days[0].get('Holidays')
    liq_travel = user_days[0].get('LiqTravel')

    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            database.update_user_days(extra_hours, holidays, liq_travel, email)
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    return jsonify({'state': 'SUCCESS'}), 200

    
#DELETE THIS if hr head count is deleted definitively
@hr_blue_print.route('/get-hr-head-count-CUE-report', methods=['POST'])
@token_required
def get_head_count_CUE_report(current_user):
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    month_input = request.form.get('month')
    year_input = request.form.get('year')
    year_short = str(year_input)[-2:]
    # logging.log(msg=month_input, level=logging.DEBUG)
    # logging.log(msg=year_input, level=logging.DEBUG)
    mes = f"{month_input}-{year_short}"
    hojas_a_eliminar = ['Dpt. FC', 'Dpt. FC ID', 'Dpt. Real', 'Dpt. Real ID', 'Dpt. Diff', 'ID Dpt. Diff', 'FC Overview']
    workbook = load_workbook(filename=file)
    for hoja in hojas_a_eliminar:
        if hoja in workbook.sheetnames:
            sheet = workbook[hoja]
            workbook.remove(sheet)
            
    workbook_mod = hiring.procesar_excel(workbook, mes)
    workbook_mod = hiring.insertar_columna_sum(workbook_mod, 'Dpt. Diff', mes)
    workbook_mod = hiring.insertar_columna_sum(workbook_mod, 'ID Dpt. Diff', mes)

    # Actualizar los totales para la hoja 'Hiring Plan 24-25 S.L.U'
    hiring.actualizar_totales(workbook_mod, 'Hiring Plan 24-25 S.L.U', 'S', 'BX', 'BY', 'BZ')

    # Actualizar los totales para la hoja 'Hiring Real'
    hiring.actualizar_totales(workbook_mod, 'Hiring Real', 'E', 'BJ', 'BK', 'BL')
    hiring.modificar_estilos(workbook_mod, 'FC Overview')

    # workbook_mod.save('Recruiting Planning Spain 24-25 Forecast vs Real_260624_SCA.xlsx')
    output = BytesIO()
    workbook_mod.save(output)
    output.seek(0)
    return send_file(output, download_name="Placeholder.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@hr_blue_print.route('/get-all-users-simple-list', methods=["GET"])
@token_required
def get_users_simple_list(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            users = database.get_users_name_mail()
            users_json = [{'FullName':data[0], 'Email':data[1], 'Department':data[2], 'City':data[3]} for data in users]
            
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    return jsonify(users_json), 200


@hr_blue_print.route('/get-user-days-buckets', methods=["POST"])
@token_required
def get_user_days_buckets(current_user):
    data = request.get_json() 
    email = data.get('email') if data.get('Email') is not None else current_user
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            user_days_data = database.get_user_days(email)
            user_days_json = [
                              {'BucketName':'ExtraHours', 'BucketValue':user_days_data[1]},
                              {'BucketName':'Holidays', 'BucketValue':user_days_data[2]},
                              {'BucketName':'LiqTravel', 'BucketValue':user_days_data[3]}  
                             ]
            return jsonify(user_days_json), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    
    
#TODO: test_get_all_managers
@hr_blue_print.route('/get-all-managers', methods=["GET"])
@token_required
def get_all_managers(current_user):
    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            managers_raw = database.get_all_managers()
            managers = [] 
            for manager in managers_raw:    
                managers.append(manager[0])
        return jsonify({'Managers': managers}), 200
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500


@hr_blue_print.route('/get-reporting-home-office', methods=['POST'])
@token_required
def create_home_office_report(current_user):
    data = request.get_json()
    starting_date = data.get('starting_date')
    finishing_date = data.get('finishing_date')
    
    try:
        if datetime.strptime(starting_date, '%Y-%m-%d') > datetime.strptime(finishing_date,'%Y-%m-%d'): 
            return jsonify({'state': 'Invalid date range'}), 400
        
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            filtered_ho_days = database.get_all_home_office_grouped(starting_date, finishing_date)
        data_dict = {}
        
        for row in filtered_ho_days:
            #set names for each list item so they correspond to row as their labels
            city, department, email, date, type_location = row
            
            if city not in data_dict:
                data_dict[city] = {}
            
            if department not in data_dict[city]:
                data_dict[city][department] = {}
                
            if email not in data_dict[city][department]:
                data_dict[city][department][email] = []
            
            data_dict[city][department][email].append({
                'Date': date,
                'Type': type_location
            })
            
        print(data_dict)
        
        #In case we need something else as output
        type_map = {
            'OFFICE': 'OFFICE',
            'PARTIAL': 'PARTIAL',
            'FULL': 'FULL'
        }

        # Check if we want to use as key the data from the Db or the value from type_map dictionary instead.
        type_colors = {
            'OFFICE': PatternFill(start_color="7954AF", end_color="7954AF", fill_type="solid"),
            'PARTIAL': PatternFill(start_color="49B5E8", end_color="49B5E8", fill_type="solid"),
            'FULL': PatternFill(start_color="2E73C0", end_color="2E73C0", fill_type="solid")
        }
        fill_date = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_city = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        fill_department = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

        # Generate a list of dates between the start and end dates
        start_date_dt = pd.to_datetime(starting_date)
        end_date_dt = pd.to_datetime(finishing_date)
        date_range = pd.date_range(start_date_dt, end_date_dt).strftime('%Y-%m-%d').tolist()
        
        logging.debug(date_range)
        
        wb = Workbook()
        ws = wb.active
        
        column_names = 1
        starting_row_names = 3
        starting_row_dates = 2
        starting_col_dates = 2
        date_col_width = 10
        
        
        for col, date in enumerate(date_range, start=starting_col_dates):
            current_cell = ws.cell(row=starting_row_dates, column=col, value=date)
            current_cell.fill = fill_date
        for col in range(starting_col_dates, starting_col_dates + len(date_range)):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = date_col_width
        
        row = starting_row_names
        ws.column_dimensions[get_column_letter(column_names)].width = 50
        for city, departments in data_dict.items():
            ws.cell(row=row, column=column_names, value=city)
            for col in range(column_names, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_city
            row+=1
            
            for department, users in departments.items():
                ws.cell(row=row, column=column_names, value=department)
                for col in range(column_names, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill_department
                row+=1
                
                for email in users:
                    ho_values = users[email]
                    for ho in ho_values:
                        ws.cell(row=row, column=column_names, value=email)
                        for col, date in enumerate(date_range, start=starting_col_dates):
                            date_from_range = pd.to_datetime(date)
                            date_from_user = pd.to_datetime(ho['Date'])
                            if date_from_user == date_from_range:
                                type_abbr = type_map[ho['Type']]
                                current_cel = ws.cell(row=row, column=col, value=type_abbr)
                                current_cel.fill = type_colors[ho['Type']]
                    row+=1
                row+=1

        # Send the file as a response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name="HomeOffice.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'), 200 
    except data_utils.pymssql.Error:
        return jsonify({'state': 'DB error'}), 500
    
    
#TODO: create_vacations_filtered_report() [hr/reporting] + test_create_vacations_filtered_report
@hr_blue_print.route('/get-reporting-vacations', methods=['POST'])
@token_required
def create_vacations_filtered_report(create_user):
    data = request.get_json()
    start_date = data.get('start_date')
    end__date = data.get('end_date')
    #ADD MORE FILTERS to vacations report
    # departments:list = data.get('departments') if len(list(data.get('departments'))) > 0 else None 
    # cities:list = data.get('cities') if len(list(data.get('cities'))) > 0 else None 
    # users:list = data.get('users') if len(list(data.get('users'))) > 0 else None 
    
    try:
        if datetime.strptime(start_date, '%Y-%m-%d') > datetime.strptime(end__date,'%Y-%m-%d'): 
            return jsonify({'state': 'Invalid date range'}), 400
        # if users is not None and (departments is not None or cities is not None):
        #     return jsonify({'state': 'Can not filter by department or city if one or several users are selected'}), 400
        
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            filtered_vacation_days = database.get_vacations_data_filtered(start_date, end__date)
        data_dict = {}
        
        #BUILD filtered vacation EXCEL REPORT
    except data_utils.pymssql.Error as ex:
        return jsonify({'state': 'DB error: ' + str(ex)}), 500
    except Exception as ex:
        return jsonify({'state': 'Error: ' + str(ex)}), 500
        

#TODO: test_add_city_days
@hr_blue_print.route('/add-city-days', methods=["PUT"])
@token_required
def add_city_days(current_user):
    data = request.get_json()
    city_days:dict = data.get("cityDays")

    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            #first check if city days registries for next year already exist
            cur_city_days:list = get_city_vacation_days()[0].json
            next_year:int = datetime.now().year +1

            exists_next_year = any(item['Year'] == str(next_year) for item in cur_city_days)
            if exists_next_year:
                return jsonify({'state': f'City vacation days for {next_year} have already been set'}), 400
            
            #we add the registries in the table
            for (city, days) in city_days.items():
                database.add_city_days(next_year, city, days)
            
            #then we add those days to the holiday bucket of every employee
            for (city, days) in city_days.items():
                users = database.get_users_in_city(city)
                for user in users:
                    user_days = database.get_user_days(user[1])
                    updated_holidays = user_days[2] + Decimal(days)
                    database.update_user_days(user_days[1], updated_holidays, user_days[3], user[1])
                       
    except data_utils.pymssql.Error as ex:
        data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME)._conn.rollback()
        return jsonify({'state': 'DB error: ' +  str(ex)}), 500
    return jsonify({'state': 'SUCCESS'}), 200


#TODO: test_remove_city_days
@hr_blue_print.route('/remove-city-days', methods=["DELETE"])
@token_required
def remove_city_days(current_user):
    data = request.get_json()
    year_del = data.get("year")

    try:
        with data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME) as database:
            #first check if city days registries for next year already exist
            cur_city_days = get_city_vacation_days()[0].json

            city_days_in_year = [item for item in cur_city_days if item['Year'] == str(year_del)]
            if len(city_days_in_year) == 0:
                return jsonify({'state': f'No registries for {year_del} exist'}), 400
            
            #we remove the registries in the table
            database.remove_city_days_year(year_del)
            
            #then we substract those days to the holiday bucket of every employee
            #TODO: comprobar que el year_del sea el proximo año, si no lo es NO QUITAR NADA DE BUCKETS
            for city_day in city_days_in_year:
                 users = database.get_users_in_city(city_day['City'])
                 for user in users:
                    user_days = database.get_user_days(user[1])
                    updated_holidays = user_days[2] - Decimal(city_day['Days'])
                    database.update_user_days(user_days[1], updated_holidays, user_days[3], user[1])
                       
    except data_utils.pymssql.Error as ex:
        data_utils.Database(Config.DATABASE_URI, Config.DATABASE_USER, Config.DATABASE_PASSWORD, Config.DATABASE_NAME)._conn.rollback()
        return jsonify({'state': 'DB error: ' +  str(ex)}), 500
    return jsonify({'state': 'SUCCESS'}), 200