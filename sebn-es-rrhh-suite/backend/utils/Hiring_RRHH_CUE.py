import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Función para crear una hoja de resumen con pronóstico, real y diferencia
def create_fc_overview(workbook, fc_sheet_name, real_sheet_name, overview_sheet_name, mes):
    # Crear o seleccionar la hoja de resumen
    if overview_sheet_name in workbook.sheetnames:
        overview_sheet = workbook[overview_sheet_name]
    else:
        overview_sheet = workbook.create_sheet(overview_sheet_name)
    
    fc_sheet = workbook[fc_sheet_name]
    real_sheet = workbook[real_sheet_name]

    # Cabeceras de la nueva hoja
    headers = ['Departament']
    months = []

    # Asumimos que los encabezados de las columnas de meses son iguales en ambas hojas
    for col in range(2, fc_sheet.max_column + 1):
        cell_value = fc_sheet.cell(row=1, column=col).value
        if cell_value is not None:
            formatted_date = cell_value if isinstance(cell_value, str) else cell_value.strftime('%b-%y')
            headers.extend([f'FC {formatted_date}', f'Real {formatted_date}', f'Diff {formatted_date}'])
            months.append(formatted_date)

    overview_sheet.append(headers)

    # Estilo de los encabezados
    for cell in overview_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Crear diccionarios de totales
    fc_totals = {}
    real_totals = {}

    # Procesar hojas de origen para crear los diccionarios
    for row in range(2, fc_sheet.max_row + 1):
        department = fc_sheet.cell(row=row, column=1).value
        if department not in fc_totals:
            fc_totals[department] = {}
        for col in range(2, fc_sheet.max_column + 1):
            cell_value = fc_sheet.cell(row=1, column=col).value
            if cell_value is not None:
                if isinstance(cell_value, datetime):
                    month = cell_value.strftime('%b-%y')
                else:
                    month = str(cell_value)
                fc_totals[department][month] = fc_sheet.cell(row=row, column=col).value

    for row in range(2, real_sheet.max_row + 1):
        department = real_sheet.cell(row=row, column=1).value
        if department not in real_totals:
            real_totals[department] = {}
        for col in range(2, real_sheet.max_column + 1):
            cell_value = real_sheet.cell(row=1, column=col).value
            if cell_value is not None:
                if isinstance(cell_value, datetime):
                    month = cell_value.strftime('%b-%y')
                else:
                    month = str(cell_value)
                real_totals[department][month] = real_sheet.cell(row=row, column=col).value

    # Añadir datos a la hoja de resumen
    all_departments = sorted(set(fc_totals.keys()).union(real_totals.keys()))

    for department in all_departments:
        row_data = [department]

        for month in months:
            fc_value = fc_totals.get(department, {}).get(month, 0)
            real_value = real_totals.get(department, {}).get(month, 0)
            
            # Convertir a números, manejando celdas vacías y cadenas
            try:
                fc_value = float(fc_value) if fc_value not in [None, ''] else 0
                real_value = float(real_value) if real_value not in [None, ''] else 0
            except ValueError:
                fc_value = 0
                real_value = 0

            diff_value = real_value - fc_value
            row_data.extend([fc_value, real_value, diff_value])

        overview_sheet.append(row_data)

    # Insertar la columna SUM después del mes especificado
    mes_col = None
    for col in range(2, overview_sheet.max_column + 1):
        cell_value = overview_sheet.cell(row=1, column=col).value
        if isinstance(cell_value, str) and f'Diff {mes}' in cell_value:
            mes_col = col
            break

    if mes_col is not None:
        mes_col += 1
        overview_sheet.insert_cols(mes_col)
        overview_sheet.cell(row=1, column=mes_col).value = 'Sum'
        overview_sheet.cell(row=1, column=mes_col).font = Font(bold=True)
        overview_sheet.cell(row=1, column=mes_col).alignment = Alignment(horizontal='center')

        # Calcula la suma de las diferencias hasta el mes especificado
        for row in range(2, overview_sheet.max_row + 1):
            sum_value = sum(overview_sheet.cell(row=row, column=col).value or 0 for col in range(2, mes_col) if 'Diff' in overview_sheet.cell(row=1, column=col).value)
            sum_cell = overview_sheet.cell(row=row, column=mes_col)
            sum_cell.value = sum_value
            sum_cell.alignment = Alignment(horizontal='center')

            # Aplica relleno rojo a los valores negativos
            if sum_value < 0:
                sum_cell.fill = red_fill

    # Estilo de las celdas
    for row in overview_sheet.iter_rows(min_row=2, max_row=overview_sheet.max_row, min_col=2, max_col=overview_sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.fill = red_fill

    # Estilo de las celdas de la primera columna
    for cell in overview_sheet['A']:
        cell.font = Font(bold=True)

# Función para procesar una hoja de cálculo
def procesar_hoja(sheet, start_col, end_col, department_col, department_id_col):
    department_totals = {}
    department_id_totals = {}
    
    # Itera sobre las filas de la hoja
    for row in range(2, sheet.max_row + 1):
        if sheet[f'A{row}'].value == "Martorell":
            continue  # Saltar filas con "Martorell" en la columna A
        
        department = sheet[f'{department_col}{row}'].value
        department_id = sheet[f'{department_id_col}{row}'].value
        
        # Inicializa los totales si no existen
        if department not in department_totals:
            department_totals[department] = [0] * (end_col - start_col + 1)
        if department_id not in department_id_totals:
            department_id_totals[department_id] = [0] * (end_col - start_col + 1)
        
        # Suma los valores por departamento e ID de departamento
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            font = cell.font
            if isinstance(value, (int, float)) and not font.strike:
                month_index = col - start_col
                department_totals[department][month_index] += value
                department_id_totals[department_id][month_index] += value
    
    return department_totals, department_id_totals

# Función para crear una hoja de resumen
def create_overview_sheet(workbook, sheet_name, headers, data_dict):
    if sheet_name in workbook.sheetnames:
        overview_sheet = workbook[sheet_name]
    else:
        overview_sheet = workbook.create_sheet(sheet_name)
    
    overview_sheet.append(headers)
    
    # Estilo de los encabezados
    for cell in overview_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    rows = []
    for key, totals in data_dict.items():
        if key is not None:
            row = [key] + [total if total != 0 else '' for total in totals]
            rows.append(row)
    
    rows.sort(key=lambda x: x[0])
    
    # Añade las filas a la hoja
    for row in rows:
        overview_sheet.append(row)
    
    # Estiliza la primera columna
    for row in overview_sheet.iter_rows(min_row=1, max_row=overview_sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)
    
    # Estiliza las celdas restantes
    for row in overview_sheet.iter_rows(min_row=1, max_row=overview_sheet.max_row, min_col=2, max_col=overview_sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    
    # Calcula y añade la fila de totales
    total_row = ['Total']
    for col in range(2, overview_sheet.max_column + 1):
        total = 0
        for row in range(2, overview_sheet.max_row + 1):
            cell_value = overview_sheet.cell(row=row, column=col).value
            try:
                total += float(cell_value or 0)
            except ValueError:
                pass
        total_row.append(total if total != 0 else '')
    
    overview_sheet.append(total_row)
    for cell in overview_sheet[overview_sheet.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    overview_sheet.column_dimensions['A'].width = 220 * 0.14 

# Función para crear una hoja de diferencias
def create_diff_sheet(workbook, sheet_name, headers, forecast_data, real_data):
    if sheet_name in workbook.sheetnames:
        diff_sheet = workbook[sheet_name]
    else:
        diff_sheet = workbook.create_sheet(sheet_name)
    
    diff_sheet.append(headers)
    
    # Estiliza los encabezados
    for cell in diff_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Filtra claves None de forecast_data y real_data
    forecast_data_filtered = {k: v for k, v in forecast_data.items() if k is not None}
    real_data_filtered = {k: v for k, v in real_data.items() if k is not None}
    
    # Asegura que ambos diccionarios filtrados tengan las mismas claves
    all_keys = set(forecast_data_filtered.keys()).intersection(real_data_filtered.keys())
    
    rows = []
    for key in sorted(all_keys):
        forecast_totals = forecast_data_filtered[key]
        real_totals = real_data_filtered[key]
        diff_totals = [r - f if r - f != 0 else '' for f, r in zip(forecast_totals, real_totals)]
        row = [key] + diff_totals
        rows.append(row)
    
    # Añade las filas a la hoja
    for row in rows:
        diff_sheet.append(row)
    
    # Añade columnas de totales para 2024 y 2024-2028
    add_total_columns(diff_sheet, 2024, 2028)
    
    # Aplica estilo a las celdas con diferencias negativas
    for row in diff_sheet.iter_rows(min_row=2, max_row=diff_sheet.max_row, min_col=2, max_col=diff_sheet.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.fill = red_fill
            cell.alignment = Alignment(horizontal='center')

# Función para añadir columnas de totales anuales
def add_total_columns(sheet, start_year, end_year):
    headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    year_cols = {year: [] for year in range(start_year, end_year + 1)}

    for col_idx, header in enumerate(headers, start=1):
        if isinstance(header, str):
            try:
                header_date = datetime.strptime(header, '%b-%y')
                year = header_date.year
                if start_year <= year <= end_year:
                    year_cols[year].append(col_idx)
            except ValueError:
                pass

    total_2024_col = sheet.max_column + 1
    total_2024_2028_col = sheet.max_column + 2

    sheet.cell(row=1, column=total_2024_col).value = 'Total 2024'
    sheet.cell(row=1, column=total_2024_2028_col).value = 'Total 2024-2028'

    for row in range(2, sheet.max_row + 1):
        total_2024 = sum(sheet.cell(row=row, column=col).value or 0 for col in year_cols[2024])
        total_2024_2028 = sum(sheet.cell(row=row, column=col).value or 0 for year in range(2024, 2029) for col in year_cols[year])
        
        sheet.cell(row=row, column=total_2024_col).value = total_2024
        sheet.cell(row=row, column=total_2024_2028_col).value = total_2024_2028

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

# Función para insertar una columna de suma en una hoja
def insertar_columna_sum(workbook, sheet_name, mes):
    sheet = workbook[sheet_name]
    mes_col = None
    
    # Encuentra la columna correspondiente al mes especificado
    for col in range(2, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if isinstance(cell_value, datetime):
            formatted_date = cell_value.strftime('%b-%y')
            if formatted_date == mes:
                mes_col = col
                break
        elif isinstance(cell_value, str) and cell_value == mes:
            mes_col = col
            break
    
    if mes_col is not None:
        # Inserta una nueva columna para la suma después de la columna del mes
        mes_col += 1
        sheet.insert_cols(mes_col)
        sheet.cell(row=1, column=mes_col).value = 'Sum'
        sheet.cell(row=1, column=mes_col).font = Font(bold=True)
        sheet.cell(row=1, column=mes_col).alignment = Alignment(horizontal='center')

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'))        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Calcula la suma para cada fila y añade el valor en la nueva columna
        for row in range(2, sheet.max_row + 1):
            sum_value = sum(sheet.cell(row=row, column=col).value or 0 for col in range(2, mes_col))
            sum_cell = sheet.cell(row=row, column=mes_col)
            sum_cell.value = sum_value
            sum_cell.alignment = Alignment(horizontal='center')
            sum_cell.border = thin_border if row < sheet.max_row else thick_border

            # Aplica relleno rojo a los valores negativos
            if sum_value < 0:
                sum_cell.fill = red_fill

        # Aplica borde grueso a las celdas de la columna de suma
        for row in range(1, sheet.max_row + 1):
            sum_cell = sheet.cell(row=row, column=mes_col)
            sum_cell.border = thick_border

    return workbook

# Función principal para procesar el archivo Excel
def procesar_excel(workbook, mes):
    sheet_forecast = workbook['Hiring Plan 24-25 S.L.U']
    sheet_real = workbook['Hiring Real']
    
    # Columnas de inicio y fin para cada hoja
    start_col_forecast = openpyxl.utils.cell.column_index_from_string('S')
    end_col_forecast = openpyxl.utils.cell.column_index_from_string('BX')
    start_col_real = openpyxl.utils.cell.column_index_from_string('E')
    end_col_real = openpyxl.utils.cell.column_index_from_string('BJ')
    
    department_col = 'B'
    department_id_col = 'A'
    
    # Procesa las hojas de pronóstico y real
    department_totals_forecast, department_id_totals_forecast = procesar_hoja(sheet_forecast, start_col_forecast, end_col_forecast, department_col, department_id_col)
    department_totals_real, department_id_totals_real = procesar_hoja(sheet_real, start_col_real, end_col_real, department_col, department_id_col)
    
    headers_forecast = ['Departament']
    for col in range(start_col_forecast, end_col_forecast + 1):
        cell_value = sheet_forecast.cell(row=1, column=col).value
        if isinstance(cell_value, datetime):
            formatted_date = cell_value.strftime('%b-%y')
        else:
            formatted_date = str(cell_value)
        headers_forecast.append(formatted_date)
    
    headers_real = ['Departament']
    for col in range(start_col_real, end_col_real + 1):
        cell_value = sheet_real.cell(row=1, column=col).value
        if isinstance(cell_value, datetime):
            formatted_date = cell_value.strftime('%b-%y')
        else:
            formatted_date = str(cell_value)
        headers_real.append(formatted_date)
    
    # Crea hojas de resumen
    create_overview_sheet(workbook, 'Dpt. FC', headers_forecast, department_totals_forecast)
    create_overview_sheet(workbook, 'Dpt. FC ID', headers_forecast, department_id_totals_forecast)
    create_overview_sheet(workbook, 'Dpt. Real', headers_real, department_totals_real)
    create_overview_sheet(workbook, 'Dpt. Real ID', headers_real, department_id_totals_real)
    
    # Crea hojas de diferencias
    create_diff_sheet(workbook, 'Dpt. Diff', headers_forecast, department_totals_forecast, department_totals_real)
    create_diff_sheet(workbook, 'ID Dpt. Diff', headers_forecast, department_id_totals_forecast, department_id_totals_real)
    
    # Añadir columnas de totales a las hojas
    for sheet_name in ['Dpt. FC', 'Dpt. FC ID', 'Dpt. Real', 'Dpt. Real ID']:
        sheet = workbook[sheet_name]
        add_total_columns(sheet, 2024, 2028)
    
    # Crear la hoja de resumen FC Overview
    create_fc_overview(workbook, 'Dpt. FC ID', 'Dpt. Real ID', 'FC Overview', mes)
    
    output_filename = 'Recruiting Planning Spain 24-25 Forecast vs Real.xlsx'
    workbook.save(output_filename)
    
    return workbook

def actualizar_totales(workbook, sheet_name, start_col, end_col, total_col_2024, total_col_2024_2028):
    sheet = workbook[sheet_name]
    
    # Encuentra los índices de las columnas de inicio, fin y totales
    start_col_idx = openpyxl.utils.cell.column_index_from_string(start_col)
    end_col_idx = openpyxl.utils.cell.column_index_from_string(end_col)
    total_col_2024_idx = openpyxl.utils.cell.column_index_from_string(total_col_2024)
    total_col_2024_2028_idx = openpyxl.utils.cell.column_index_from_string(total_col_2024_2028)
    
    # Identifica las columnas que corresponden a los meses del 2024
    cols_2024 = []
    for col in range(start_col_idx, end_col_idx + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if isinstance(cell_value, datetime) and cell_value.year == 2024:
            cols_2024.append(col)
        elif isinstance(cell_value, str):
            try:
                cell_date = datetime.strptime(cell_value, '%b-%y')
                if cell_date.year == 2024:
                    cols_2024.append(col)
            except ValueError:
                continue

    # Itera sobre las filas de la hoja
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == "TOTALES":
            # Realiza la suma de toda la columna hasta la fila anterior
            total_2024 = sum(
                sheet.cell(r, total_col_2024_idx).value or 0
                for r in range(2, row)
                if isinstance(sheet.cell(r, total_col_2024_idx).value, (int, float))
            )
            total_2024_2028 = sum(
                sheet.cell(r, total_col_2024_2028_idx).value or 0
                for r in range(2, row)
                if isinstance(sheet.cell(r, total_col_2024_2028_idx).value, (int, float))
            )
            
            # Actualiza la fila de "TOTALES"
            sheet.cell(row=row, column=total_col_2024_idx).value = total_2024
            sheet.cell(row=row, column=total_col_2024_2028_idx).value = total_2024_2028
            break
        else:
            total_2024 = 0
            total_2024_2028 = 0

            # Suma los valores que no están tachados
            for col in range(start_col_idx, end_col_idx + 1):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                font = cell.font
                if isinstance(value, (int, float)) and not font.strike:
                    if col in cols_2024:
                        total_2024 += value
                    total_2024_2028 += value
            
            # Actualiza las columnas de totales
            sheet.cell(row=row, column=total_col_2024_idx).value = total_2024
            sheet.cell(row=row, column=total_col_2024_2028_idx).value = total_2024_2028

    # Estiliza las columnas de totales
    for col in [total_col_2024_idx, total_col_2024_2028_idx]:
        sheet.cell(row=1, column=col).font = Font(bold=True)
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        
def modificar_estilos(workbook, overview_sheet_name):
    overview_sheet = workbook[overview_sheet_name]
    
    # Definir bordes y colores
    # thin_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    # thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), 
    #                       top=Side(style='thick'), bottom=Side(style='thick'))
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    light_blue_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    dark_blue_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

    # Ajuste del ancho de las columnas
    for col in overview_sheet.columns:
        col_letter = col[0].column_letter
        overview_sheet.column_dimensions[col_letter].width = 100 * 0.14
    overview_sheet.column_dimensions['A'].width = 220 * 0.14

    # Mapear columnas FC
    fc_col_indices = {}
    for col_idx in range(2, overview_sheet.max_column + 1):
        header_value = overview_sheet.cell(row=1, column=col_idx).value
        if header_value.startswith('FC'):
            fc_col_indices[header_value] = col_idx

    # Aplicar colores alternados cada 3 columnas en los encabezados desde la columna B hasta la columna "Sum"
    sum_col = None
    for col_idx in range(2, overview_sheet.max_column + 1):
        cell = overview_sheet.cell(row=1, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value == 'Sum':
            cell.fill = yellow_fill
            sum_col = col_idx
            break
        if (col_idx - 2) // 3 % 2 == 0:
            cell.fill = dark_blue_fill
        else:
            cell.fill = light_blue_fill

    # Aplicar el relleno verde claro a las columnas "Real..." si su valor es igual al de la columna "FC..."
    for row in overview_sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            if cell.value == 0:
                cell.value = None
            if isinstance(cell.value, (int, float)):
                header_value = overview_sheet.cell(row=1, column=cell.column).value
                if header_value.startswith('Real'):
                    fc_header = 'FC ' + header_value.split('Real ')[1]
                    fc_col = fc_col_indices.get(fc_header)
                    if fc_col and cell.value == overview_sheet.cell(row=cell.row, column=fc_col).value:
                        cell.fill = green_fill
                elif header_value.startswith('Diff') and cell.value < 0:
                    cell.fill = red_fill

    # Aplicar relleno azul claro a la columna A desde la segunda fila hasta la fila de totales
    max_row = overview_sheet.max_row
    for row in range(2, max_row):
        overview_sheet[f'A{row}'].fill = light_blue_fill

    # Aplicar bordes externos y relleno azul oscuro solo a la celda A de la fila de totales
    for col_idx in range(1, sum_col + 1):
        cell = overview_sheet.cell(row=max_row, column=col_idx)
        if col_idx == 1:
            cell.fill = dark_blue_fill
        else:
            cell.fill = PatternFill(fill_type=None)  # Dejar el resto en blanco
        cell.border = Border(
            left=Side(style='thick') if col_idx == 1 else None,
            right=Side(style='thick') if col_idx == sum_col else None,
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        # Aplicar relleno rojo a las celdas con valores negativos en la fila total
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.fill = red_fill
    
    # Aplicar bordes externos a la columna "Sum"
    if sum_col:
        for row in range(1, overview_sheet.max_row + 1):
            overview_sheet.cell(row=row, column=sum_col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin') if row == 1 else None,
                bottom=Side(style='thin') if row == overview_sheet.max_row else None
            )

# Ejecutar la función principal
# filename = 'Recruiting Planning Spain 24-25 Forecast vs Real_260624_cristina.xlsx'
# mes = 'Jun-24'
# hojas_a_eliminar = ['Dpt. FC', 'Dpt. FC ID', 'Dpt. Real', 'Dpt. Real ID', 'Dpt. Diff', 'ID Dpt. Diff', 'FC Overview']
# workbook = openpyxl.load_workbook(filename)
# for hoja in hojas_a_eliminar:
#     if hoja in workbook.sheetnames:
#         sheet = workbook[hoja]
#         workbook.remove(sheet)
        
# workbook_mod = procesar_excel(workbook)
# workbook_mod = insertar_columna_sum(workbook_mod, 'Dpt. Diff', mes)
# workbook_mod = insertar_columna_sum(workbook_mod, 'ID Dpt. Diff', mes)

# # Actualizar los totales para la hoja 'Hiring Plan 24-25 S.L.U'
# actualizar_totales(workbook_mod, 'Hiring Plan 24-25 S.L.U', 'S', 'BX', 'BY', 'BZ')

# # Actualizar los totales para la hoja 'Hiring Real'
# actualizar_totales(workbook_mod, 'Hiring Real', 'E', 'BJ', 'BK', 'BL')
# modificar_estilos(workbook_mod, 'FC Overview')

# workbook_mod.save('Recruiting Planning Spain 24-25 Forecast vs Real_260624_SCA.xlsx')